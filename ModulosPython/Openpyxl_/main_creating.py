# openpyxl para arquivos Excel xlsx, xlsm, xltx e xltm (instalação)
# Com essa biblioteca será possível ler e escrever dados em células
# específicas, formatar células, inserir gráficos,
# criar fórmulas, adicionar imagens e outros elementos gráficos às suas
# planilhas. Ela é útil para automatizar tarefas envolvendo planilhas do
# Excel, como a criação de relatórios e análise de dados e/ou facilitando a
# manipulação de grandes quantidades de informações.
# Instalação necessária: pip install openpyxl
# Documentação: https://openpyxl.readthedocs.io/en/stable/
from pathlib import Path
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

ROOT_FOLDER = Path(__file__).parent
WORKBOOK_PATH = ROOT_FOLDER / 'workbook.xlsx'

workbook = Workbook()
workbook.create_sheet('My_sheet')
worksheet: Worksheet = workbook['My_sheet']
workbook.remove(workbook['Sheet'])

worksheet.cell(1, 1, "Nome")
worksheet.cell(1, 2, "Idade")
worksheet.cell(1, 3, "Nota")

students = [
    # nome      idade nota
    ['João',    14,   5.5],
    ['Maria',   13,   9.7],
    ['Luiz',    15,   8.8],
    ['Alberto', 16,   10],
]

for row, students_row in enumerate(students, start=2):
    for column, item in enumerate(students_row, start=1):
        print(row, column, item)
        worksheet.cell(row, column, item)

workbook.save(WORKBOOK_PATH)