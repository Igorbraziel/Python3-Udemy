# openpyxl- ler e alterar dados de uma planilha
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

ROOT_FOLDER = Path(__file__).parent
WORKBOOK_PATH = ROOT_FOLDER / 'workbook.xlsx'

workbook = load_workbook(WORKBOOK_PATH)
worksheet: Worksheet = workbook.active

for row in worksheet.iter_rows():
    for item in row:
        print(item.value, end='\t')
    print()

worksheet['B4'] = 22

workbook.save(WORKBOOK_PATH)
